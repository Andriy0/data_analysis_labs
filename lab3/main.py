import openpyxl as opl
import matplotlib.pyplot as plt
import numpy as np

wb = opl.load_workbook('./../PV2013June.xlsx')
birth_date = '07'
ws = wb[birth_date + '-06-2013']

def read_column(ws, col):
  return [ws.cell(row=i, column=col).value for i in range(5, ws.max_row + 1)]

def ecdf(data):
  x, counts = np.unique(data, return_counts=True)
  cumsum = np.cumsum(counts)
  return x, cumsum / cumsum[-1]

def plot_ecdf(data, title=''):
  x, y = ecdf(data)
  x = np.insert(x, 0, x[0])
  y = np.insert(y, 0, 0.)
  plt.plot(x, y, drawstyle='steps-post')
  plt.grid(True)
  plt.title(title)
  plt.show()

def print_percentiles(data, type=''):
  [print(f"{percent}th percentile of {type} data is: {np.percentile(data, percent)}") for percent in [25, 50, 75]]

def filter_zeros(data):
  return list(filter(lambda x: x > 0, data))

temperature = read_column(ws, 6)
print_percentiles(temperature, 'temperature')
plot_ecdf(temperature, 'temperature')

frequency = read_column(ws, 9)
print_percentiles(frequency, 'frequency')
# plot_ecdf(frequency, 'frequency')

generated_power = read_column(ws, 16)
print_percentiles(generated_power, 'generated power')
# plot_ecdf(generated_power, 'generated_power')
