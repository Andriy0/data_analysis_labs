import openpyxl as opl
import matplotlib.pyplot as plt
import numpy as np
from scipy import stats
import math

ws = opl.load_workbook('./PV2013June.xlsx')['07-06-2013']

def read_column(ws, col):
  return [ws.cell(row=i, column=col).value for i in range(5, ws.max_row + 1)]

def filter_zeros(data):
  return list(filter(lambda x: x > 0, data))

def RawMoment(xs, k):
    return sum(x**k for x in xs) / len(xs)

def find_middle(lst):
    length = len(lst)  # Get the length of the list
    middle_index = length // 2
    return lst[middle_index]

def CentralMoment(xs, k):
    mean = RawMoment(xs, 1)
    return sum((x - mean)**k for x in xs) / len(xs)

def StandardizedMoment(xs, k):
    var = CentralMoment(xs, 2)
    std = math.sqrt(var)
    return CentralMoment(xs, k) / std**k

def Skewness(xs):
    return StandardizedMoment(xs, 3)

def ecdf(a):
    x, counts = np.unique(a, return_counts=True)
    cusum = np.cumsum(counts)
    return x, cusum / cusum[-1]  
   
def Median(xs):
    cdf = ecdf(xs)
    return find_middle(cdf[0])

def PearsonMedianSkewness(xs):
    median = Median(xs)
    mean = RawMoment(xs, 1)
    var = CentralMoment(xs, 2)
    std = math.sqrt(var)
    gp = 3 * (mean - median) / std
    return gp

dataArray1, title = filter_zeros(read_column(ws, 6)), 'Temperature'
# dataArray1, title = filter_zeros(read_column(ws, 9)), 'Frequency'
# dataArray1, title = filter_zeros(read_column(ws, 16)), 'Generated Power'

mu, sigma = np.mean(dataArray1), np.std(dataArray1) # mean and standard deviation
dataArray2 = np.random.normal(mu, sigma, len(dataArray1))
eval_points = np.linspace(np.min(dataArray1), np.max(dataArray1))
kde1=stats.gaussian_kde(dataArray1)
y1 = kde1.pdf(eval_points)
kde2=stats.gaussian_kde(dataArray2)
y2 = kde2.pdf(eval_points)
plt.plot(eval_points, y1)
plt.plot(eval_points, y2)
plt.title(title)
plt.grid(True)
plt.show()

print('For original', title, 'data')
print('Raw Moment:', RawMoment(dataArray1, 1))
print('Central Moment:', CentralMoment(dataArray1, 2))
print('Median:', Median(dataArray1))
print('Skewness:', Skewness(dataArray1))
print('Pearson Median Skewness:', PearsonMedianSkewness(dataArray1))
print('\n')
print('For modelled', title, 'data')
print('Raw Moment:', RawMoment(dataArray2, 1))
print('Central Moment:', CentralMoment(dataArray2, 2))
print('Median:', Median(dataArray2))
print('Skewness:', Skewness(dataArray2))
print('Pearson Median Skewness:', PearsonMedianSkewness(dataArray2))
