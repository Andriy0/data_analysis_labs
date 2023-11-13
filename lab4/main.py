import openpyxl as opl
import matplotlib.pyplot as plt
import numpy as np

ws = opl.load_workbook('./../PV2013June.xlsx')['07-06-2013']

def read_column(ws, col):
  return [ws.cell(row=i, column=col).value for i in range(5, ws.max_row + 1)]

def ecdf(data):
  x, counts = np.unique(data, return_counts=True)
  cumsum = np.cumsum(counts)
  return x, cumsum / cumsum[-1]

def plot_ecdf(a,b,title=''):
  x1, y1 = ecdf(a)
  x1 = np.insert(x1, 0, x1[0])
  y1 = np.insert(y1, 0, 0.)
  x2, y2 = ecdf(b)
  x2 = np.insert(x2, 0, x2[0])
  y2 = np.insert(y2, 0, 0.)
  plt.plot(x1, y1, drawstyle='steps-post')
  plt.plot(x2, y2, drawstyle='steps-post')
  plt.title(title)
  plt.grid(True)
  plt.show()

def filter_zeros(data):
  return list(filter(lambda x: x > 0, data))

data, title = read_column(ws, 6), 'temperature'
# data, title = read_column(ws, 9), 'frequency'
# data, title = read_column(ws, 16), 'generated power'
mu, sigma = np.mean(data), np.std(data)
normal_distribution = np.random.normal(mu, sigma, len(data))
plot_ecdf(ecdf(data), ecdf(normal_distribution), title)
