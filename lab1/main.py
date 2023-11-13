import openpyxl as opl
from statistics import mean
import matplotlib.pyplot as plt

wb = opl.load_workbook('./../PV2013June.xlsx')
birth_date = '07'
ws = wb[birth_date + '-06-2013']
# row_names = [ws.cell(row=3, column=i).value for i in range(2, ws.max_column + 1)]
# print(row_names)
data = [ws.cell(row=i, column=3).value for i in range(5, ws.max_row + 1)]
# print(data)

def time_from_minutes(time_in_mins):
  return '{:02d}:{:02d}'.format(*divmod(time_in_mins, 60))

def radiation_start_time(data):
  for i in range(len(data)):
    if data[i] > 0:
      return time_from_minutes(i * 5)
      
def radiation_end_time(data):
  data = data[::-1]
  for i in range(len(data)):
    if data[i] > 0:      
      return time_from_minutes((23 * 60 + 55) - i * 5)

print('Maximum value:', max(data))
data_without_zeros = list(filter(lambda x: x > 0, data))
print('Minimum value:', min(data_without_zeros))
print('Average value for the day:', mean(data))
print('Average value when sensor was receiving radiation:', mean(data_without_zeros))
print('Radiation start time:', radiation_start_time(data))
print('Radiation end time:', radiation_end_time(data))
print('Solar day duration:', time_from_minutes(len(data_without_zeros) * 5))
print('Solar day duration in %: {:.03f}'.format(len(data_without_zeros) * 5 * 100 / (23 * 60 + 55)))

def plotData(data):
  plt.plot(range(0,len(data)), data)
  plt.title('Solar Radiation Graph')
  plt.show()

plotData(data)
