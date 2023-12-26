import openpyxl as opl
from scipy.stats import pearsonr

ws = opl.load_workbook('./PV2013June.xlsx')['07-06-2013']

def read_column(ws, col):
  return [ws.cell(row=i, column=col).value for i in range(5, ws.max_row + 1)]

def print_stats(dataArray1, dataArray2, title1, title2):
  print(title1, 'and', title2)
  print(pearsonr(dataArray1, dataArray2))
  print('\n')

data = [[read_column(ws, 6), 'Temperature'],
        [read_column(ws, 9), 'Frequency'],
        [read_column(ws, 16), 'Generated Power']]

for i in range(3):
  print_stats(data[i-1][0], data[i][0], data[i-1][1], data[i][1])
