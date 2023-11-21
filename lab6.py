import openpyxl as opl
import numpy as np
from scipy import stats

ws = opl.load_workbook('./PV2013June.xlsx')['07-06-2013']

def read_column(ws, col):
  return [ws.cell(row=i, column=col).value for i in range(5, ws.max_row + 1)]

dataArray1, title1 = read_column(ws, 6), 'Temperature'
# dataArray1, title1 = read_column(ws, 9), 'Frequency'
# dataArray1, title1 = read_column(ws, 16), 'Generated Power'

mu, sigma = np.mean(dataArray1), np.std(dataArray1) # mean and standard deviation
dataArray2 = np.random.normal(mu, sigma, len(dataArray1))

print('Between original and modelled', title1)
print('Correlation:')
print(np.corrcoef(dataArray1, dataArray2))
print('Covariance:')
print(np.cov(dataArray1, dataArray2))
print('Pearson Correlation:')
print(stats.pearsonr(dataArray1, dataArray2))
print('\n')

# dataArray2, title2 = read_column(ws, 6), 'Temperature'
dataArray2, title2 = read_column(ws, 9), 'Frequency'
# dataArray2, title2 = read_column(ws, 16), 'Generated Power'

print('Between', title1, 'and', title2)
print('Correlation:')
print(np.corrcoef(dataArray1, dataArray2))
print('Covariance:')
print(np.cov(dataArray1, dataArray2))
print('Pearson Correlation:')
print(stats.pearsonr(dataArray1, dataArray2))
