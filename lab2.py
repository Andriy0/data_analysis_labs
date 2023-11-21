import openpyxl as opl
import matplotlib.pyplot as plt
import numpy as np

wb = opl.load_workbook('./PV2013June.xlsx')
birth_date = '07'
ws = wb[birth_date + '-06-2013']

def read_column(ws, col):
  return [ws.cell(row=i, column=col).value for i in range(5, ws.max_row + 1)]

def plotData(data, title='', xlabel='', ylabel=''):
  plt.plot(range(0,len(data)), data)
  plt.title(title)
  plt.xlabel(xlabel)
  plt.ylabel(ylabel)
  plt.show()

def plotHistogram(data, bins, title='', xlabel='', ylabel=''):
  plt.hist(data,bins)
  plt.title(title)
  plt.xlabel(xlabel)
  plt.ylabel(ylabel)
  plt.show()

def plotPMF(data, bins, title='', xlabel='', ylabel=''):
  heights,bins = np.histogram(data,bins)
  heights = heights/sum(heights)
  plt.bar(bins[:-1],heights,width=(max(bins) - min(bins))/len(bins), color="blue", alpha=0.5)
  plt.title(title)
  plt.xlabel(xlabel)
  plt.ylabel(ylabel)
  plt.show()

def filter_zeros(data):
  return list(filter(lambda x: x > 0, data))

temperature = read_column(ws, 6)
# print(temperature)
# plotData(temperature, 'Temperature Line Graph', 'n', 'temperature')
# plotHistogram(temperature, 100, 'Temperature Histogram', 'temperature', 'frequency (occurences)')
# plotPMF(temperature, 100, 'Temperature PMF', 'temperature', 'probability')

frequency = read_column(ws, 9)
# print(frequency)
# plotHistogram(filter_zeros(frequency), 3, 'Frequency Histogram', 'frequency', 'frequency (occurences)')
# plotData(frequency, 'Frequency Line Graph', 'n', 'frequency')
# plotPMF(filter_zeros(frequency), 3, 'Frequency PMF', 'frequency', 'probability')

generated_power = read_column(ws, 16)
# print(generated_power)
# plotData(generated_power, 'Generated Power Line Graph', 'n', 'generated power')
# plotHistogram(filter_zeros(generated_power), 29, 'Generated Power Histogram', 'generated power', 'frequency (occurences)')
plotPMF(filter_zeros(generated_power), 29, 'Generated Power PMF', 'generated power', 'probability')
