import openpyxl as opl
import scipy
from scipy.stats import shapiro

ws = opl.load_workbook('./PV2013June.xlsx')['07-06-2013']

def read_column(ws, col):
  return [ws.cell(row=i, column=col).value for i in range(5, ws.max_row + 1)]

def filter_zeros(data):
  return list(filter(lambda x: x > 0, data))

def print_stats(dataArray, title):
  stat, p = shapiro(dataArray)
  print(title)
  print('Statistics=%.3f, p=%.3f' % (stat, p))
  # interpret results
  alpha = 0.05
  if p > alpha:
      print('Sample looks Gaussian (fail to reject H0)')
  else:
      print('Sample does not look Gaussian (reject H0)')
  res=scipy.stats.anderson(dataArray)
  print(res.statistic)
  print(res.critical_values)
  print(res.significance_level)
  print('\n')

data = [[filter_zeros(read_column(ws, 6)), 'Temperature'],
        [filter_zeros(read_column(ws, 9)), 'Frequency'],
        [filter_zeros(read_column(ws, 16)), 'Generated Power']]

for data_set in data:
  print_stats(data_set[0], data_set[1])
